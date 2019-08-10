from urllib.request import urlopen
import urllib.parse
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
import xlrd
import pandas as pd
import time, random

# 카테고리별 시트 준비
def setup_sheet(category):
    sheet = file.create_sheet(category)

    # header 입력
    sheet['B1'].value = '상품명'
    sheet['C1'].value = '개수'
    sheet['D1'].value = '재고'

    # header 스타일 지정
    font = Font(size=12, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    fill_lightgray = PatternFill(patternType='solid', fgColor=Color('D5D5D5'))
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

    for index in range(2, 5):
        sheet.cell(row=1, column=index).font = font
        sheet.cell(row=1, column=index).alignment = align_center
        sheet.cell(row=1, column=index).fill = fill_lightgray
        sheet.cell(row=1, column=index).border = border_thin

# 메뉴 링크 뽑아오기
def extract_c_link():
    html = urlopen("http://www.sticker-kl.de/")
    soup = BeautifulSoup(html, "html.parser")

    category = ['']
    category = ['Activity Sticker', 'Scratch & Sniff', 'Sticker Sortimente', 'Feiern', 'Glow in the Dark (Leuchtsticker)', 'Halloween', 'Fasching', 'Weihnachten', 'Ostern', 'Diverse']
    c_link = {}

    for code in soup.select('ul.rh-vmenu > li > a'):
        for thing in category:
            if thing in code:
                setup_sheet(thing)
                c_link[thing] = code

    # 카테고리 중 Tiere의 하위 메뉴의 링크가 포함된 코드 추출
    html = urlopen('https://sticker-kl.de/tiere-c-22.html?bigwareCsid=rn7alpb1i7k9cu9vku2913q730')
    soup = BeautifulSoup(html, "html.parser")

    tiere = ['Bären', 'Dinosaurier', 'Hunde', 'Katzen']

    for code in soup.select('ul.rh-vmenu > li > ul.rh-vmenu > li > a'):
        for thing in tiere:
            if thing in code:
                setup_sheet(thing)
                c_link[thing] = code
    
    # 카테고리 링크만 추출
    for key, value in c_link.items():
        if 'href' in value.attrs:
            c_link[key] = value.attrs['href']
    
    # 'Tiere'의 경우(카테고리 링크 추출 다음에 놓기)
    setup_sheet('Tiere')
    c_link['Tiere'] = 'https://sticker-kl.de/tiere-c-22.html?bigwareCsid=rn7alpb1i7k9cu9vku2913q730'

    return c_link

# 아이템 링크 추출하기
def extract_i_link(c_link):
    page_1 = urlopen(c_link)
    soup = BeautifulSoup(page_1, "html.parser")

    i_link = []

    for link in soup.select('td.itemListing-data > a'):
        if 'class' not in link.attrs and 'href' in link.attrs:
            i_link.append(link.attrs['href'])

    # 한 페이지로 끝나지 않는 경우
    right = soup.find('td', {'class':'smallText', 'align':'right'})
    right = right.find_all('a')

    if right: # 페이지가 두 페이지 이상이면
        pages = []
        # 페이지 링크 추출하여 저장
        for link in right:
            pages.append(link.get('href'))
        # 마지막 링크는 다음 페이지 링크이므로 삭제(2페이지 링크가 또 나옴)
        del pages[-1]
        # 각 페이지 마다 아이템 링크 추출
        for page in pages:
            html = urlopen(page)
            soup = BeautifulSoup(html, "html.parser")
            for link in soup.select('td.itemListing-data > a'):
                if 'class' not in link.attrs and 'href' in link.attrs:
                    i_link.append(link.attrs['href'])

    items = []
    
    # 아스키 코드 에러 방지
    for link in i_link:
        parts = urllib.parse.urlsplit(link)
        parts = list(parts)
        parts[2] = urllib.parse.quote(parts[2])
        html = urllib.parse.urlunsplit(parts)
        items.append(html)
        
    return items

# 아이템 정보를 액셀에 저장
def save_info(category, num, i_link):
    page = urlopen(i_link)
    soup = BeautifulSoup(page, "html.parser")

    # 아이템 정보 추출
    title = soup.select('td.pageHeading > h1 > b > h1')[0].text
    amount = soup.select('span.smallText')[0].text
    stock = int(soup.find_all('option')[-1].text)

    # 아이템 정보 입력
    sheet = file[category]
    sheet.cell(row=num, column=2).value = title
    sheet.cell(row=num, column=3).value = amount
    sheet.cell(row=num, column=4).value = stock

    # 재고량 별로 색상 지정
    if stock <= 15:
        font = Font(size=11, bold=True, color='FF0000') # 빨강
    elif stock <= 30:
        font = Font(size=11, bold=True, color='FF9100') # 주황
    elif stock <= 50:
        font = Font(size=11, bold=True, color='52E252') # 초록
    elif stock <= 100:
        font = Font(size=11, bold=True, color='28A0FF') # 파랑
    elif stock <= 150:
        font = Font(size=11, bold=True, color='96A5FF') # 보라
    else:
        font = Font(size=11)

    align_center = Alignment(horizontal='center', vertical='center')

    sheet.cell(row=num, column=4).font = font
    sheet.cell(row=num, column=4).alignment = align_center

def finish(category):
    sheet = file[category]
    
    # 셀 너비 조정(상품명, 개수)
    sheet.column_dimensions['B'].width = 50 
    sheet.column_dimensions['C'].width = 17

    file.save('E://Desktop//sticker_kl.xlsx')
  
file = openpyxl.Workbook()
# file = openpyxl.load_workbook('E://Desktop//sticker_kl.xlsx')

c_link = extract_c_link()

for category, link in c_link.items():
    i_link = extract_i_link(link)
    n = 2
    for item in i_link:
        save_info(category, n, item)
        n = n + 1
        time.sleep(1 + random.randrange(0, 20))
    finish(category)

file.remove(file['Sheet'])

file.save('E://Desktop//sticker_kl.xlsx')
